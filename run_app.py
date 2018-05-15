#!/usr/bin/env python
# -*- coding: utf-8 -*-

from __future__ import division
from gooey import Gooey, GooeyParser
import pandas as pd
from math import ceil
from openpyxl import load_workbook
from pandas import ExcelWriter
import warnings
import os, shutil

@Gooey(language='russian', dump_build_config=True, program_name="Подготовка файлов для МСХ")

def main():

    """Executing script"""

    desc = u'Выбериате вариант и нажмите на кнопку "Запуск"'

    my_cool_parser = GooeyParser(description=desc)

    verbosity = my_cool_parser.add_mutually_exclusive_group()
    verbosity.add_argument('-t', '--verbozze', dest='First', action="store_true", help="Сформировать общий файл и вручную проставить порядковый номер")
    verbosity.add_argument('-q', '--quiet', dest='Second', action="store_true", help="Разбить на файлы (максимум 20 записей)")
    verbosity.add_argument('-e', '--extra', dest='Third', action="store_true", help="Разбить на файлы по субъектам (максимум 20 записей)")

    global args

    args = my_cool_parser.parse_args()
    global input_path
    global output_path
    global template_all_path
    global template_short_path
    global example_path

    input_path = u'//a104124/A/input/'
    output_path = u'//a104124/A/output/'
    template_all_path = u'//a104124/A/example/template_full.xlsx'
    template_short_path = u'//a104124/A/example/template_short.xlsx'
    example_path = u'//a104124/A/example/'

    all_df = read_concat(input_path)

    if args.First ==True:
        split_save(all_df)
        print 'Общий файл сохранен. Проставьте, пожалуйста, порядковый номер, сохраните файл и нажмите "След. этап" для нарезки файлов'
    elif args.Second == True:
        print 'second'
        temp = read_index()
        split_save_short (all_df)
        split_save_invest (all_df)
        print ('Готово.')
        print ("""Результат находится в папке: \\\\a104124\A\output""")
    elif args.Third == True:
        temp = read_index()
        all_df.loc[all_df['level_8']==u'Московская обл.', 'level_8'] = u'Московская область'
        for region in all_df['level_8'].unique():
            split_save_region(all_df, region)

    else:
        print ('На прошлом шаге не был выбран вариант, нажмите на кнопку редактировать')

import warnings
warnings.filterwarnings("ignore")

def clear_folder ():

    """If folder is not empty all files are removing"""

    folder = output_path
    for the_file in os.listdir(folder):
        file_path = os.path.join(folder, the_file)
        try:
            if os.path.isfile(file_path):
                os.unlink(file_path)
        except Exception as e:
            print(e)

def read_concat (input_path):

    """Reads data from public folder into full dataset"""

    count = 0
    all_df = pd.DataFrame()
    for i in os.listdir(input_path):
        rees = pd.read_excel(input_path + u'{}'.format(i), sheetname = u'Прил 2')
        rees= rees.reset_index()
        ind = []
        for index_name in rees[rees['level_0']==u'Итого:'].index.astype(int):
            ind.append(index_name)
        rees = rees.iloc[ind[0]+1:,:]
        rees = rees[~rees['level_3'].isnull()]
        rees = rees[rees['level_3'] != '']
        count+=rees.shape[0]
        print 'длина полного файла составляет {} '.format(count)
        rees['file_name'] = i
        rees =rees.reset_index(drop=True)
        rees['raw_number'] = rees.index
        all_df = pd.concat([all_df, rees], axis=0)
    all_df['uid_first'] = all_df['level_3'].astype(unicode) + all_df['level_11'].astype(unicode) + all_df['level_12'].astype(unicode) + all_df['level_17'].astype(unicode) + all_df['level_19'].astype(unicode) + all_df['level_21'].astype(str) + all_df['level_23'].fillna(0).astype(int).astype(unicode) + all_df['file_name'].astype(unicode) + all_df['raw_number'].astype(unicode)
    try:
        manual_index = pd.read_json (example_path + 'index.json')
        print ('всего в справочнике индексов {}'.format(manual_index.shape[0]))
        all_df = pd.merge(all_df, manual_index, left_on='uid_first', right_on='uid', how='left')
        all_df = all_df.drop_duplicates(subset='uid_first', keep='first')
        all_df['level_1'] = all_df['true_index']

    except:
        pass

    return all_df

def correct_percent (col):

    """Correcting interest rate of format is incorrect"""


    if (col > 1) and (type(col) == float):
        try:
            return col/100
        except:
            return col
    else:
        return col

def save_all(all_df):

    """Concating all into one file"""

    book = load_workbook(template_all_path)
    writer = ExcelWriter(output_path + u'Реестр_полный_ {} записей.xlsx'.format(all_df.shape[0]), engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    try:
        del all_df['uid_first']
        del all_df['uid']
        del all_df['true_index']
    except:
        pass
    all_df.to_excel(writer, u'Прил 2', startrow = 26, header=False, index=False)
    writer.save()

def save_cut(cut_df, i, name):

    """Cutting full file into parts according to the lengh"""

    book = load_workbook(template_short_path)
    writer = ExcelWriter(output_path + u'Реестр'+ name+ u'{} часть_'.format(i+1) + '{}'.format(cut_df.shape[0]) +  u' записей.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    cut_df.to_excel(writer, u'Прил 2', startrow = 26, header=False, index=False)
    writer.save()

def read_index ():

    """Reading extra index"""
    
    for i in os.listdir(output_path):
        if i.startswith(u'Реестр_полный_'):

            rees = pd.read_excel(output_path + i)
            print 'всего считано записей {}'.format(rees.shape[0])
            rees= rees.reset_index()
            ind = []
            for i in rees[rees['level_0']==u'Итого:'].index.astype(int):
                ind.append(i)
            rees = rees.iloc[ind[0]+1:,:]
            rees = rees[~rees['level_3'].isnull()]
            rees = rees[rees['level_3'] != '']
            print 'всего записей без шапки {}'.format(rees.shape[0])

            rees['uid'] = rees['level_3'].astype(unicode) + rees['level_11'].astype(unicode) + rees['level_12'].astype(unicode) + rees['level_17'].astype(unicode) + rees['level_19'].astype(unicode) + rees['level_21'].astype(str) + rees['level_23'].fillna(0).astype(int).astype(unicode) + rees['Unnamed: 7'].astype(unicode) + rees['Unnamed: 8'].astype(unicode)

            not_empty = rees[~rees['level_1'].isnull()]

            print not_empty.shape[0]
            if not_empty.shape[0]>0:
                not_empty = not_empty.rename(columns = {'level_1':'true_index'})
                not_empty[['uid', 'true_index']].to_json(example_path + 'index.json')
                print 'количество непустых {}'.format(not_empty.shape[0])
            else:
                pass
    return rees

def count_num_parts (all_df):

    """Shows the number of parts to split data"""

    cut_shape_all = 0
    all_df = all_df.reset_index(drop=True)
    a = all_df.shape[0]/20
    num_parts = int(ceil(a))
    print 'всего получилось {} частей'.format(num_parts)
    return num_parts

def split_save (all_df):

    """Splits into tables with max 20 rows"""

    clear_folder ()
    all_df['level_18'] = pd.to_datetime(all_df['level_18'])
    all_df['level_18'] =all_df['level_18'].dt.date
    all_df = all_df.drop(['level_0'], axis=1)
    all_df = all_df.reset_index(drop=True)
    all_df = all_df.reset_index()
    all_df['index'] =all_df['index']+1
    all_df = all_df.rename(columns={'index':'level_0'})
    all_df['level_20'] = all_df['level_20'].apply(correct_percent)

    save_all(all_df)

def split_save_short (all_df):

    """Save shrot-term requests"""

    all_df = read_index()
    try:
        del all_df['Unnamed: 7']
        del all_df['Unnamed: 8']
        del all_df['uid']
    except:
        pass

    short_term = all_df[all_df['level_16']==u'до 1 года']

    cut_shape_all = 0
    first = 0
    last = 20

    print ' '
    print 'Краткосрочные заявки:'
    print 'Количество краткосрочных заявок составляет {}'.format(short_term.shape[0])

    num_parts = count_num_parts(short_term)

    if num_parts>0:
        for i in range(num_parts):
            print 'Краткосрочные. Сохраняем часть {}'.format(i+1)
            cut_df = short_term[first:last]
            cut_df = cut_df.drop(['level_0'], axis=1)
            cut_df = cut_df.reset_index(drop=True)
            cut_df = cut_df.reset_index()
            cut_df['index'] =cut_df['index']+1
            cut_shape = cut_df.shape[0]
            cut_shape_all+=cut_shape
            name = u'_краткосрочные_'
            save_cut(cut_df, i, name)
            print 'Краткосрочные. длина сохраненного файла {} записей'.format(cut_shape)
            first+=20
            last+=20

    print 'Краткосрочные. всего сохранили {} записей'.format(cut_shape_all)

def split_save_invest (all_df):

    """Save long-term requests"""

    all_df = read_index()
    try:
        del all_df['Unnamed: 7']
        del all_df['Unnamed: 8']
        del all_df['uid']
    except:
        pass
        
    long_term = all_df[all_df['level_16']!=u'до 1 года']

    cut_shape_all = 0
    first = 0
    last = 20

    print ' '
    print 'Инвестиционные заявки:'
    print 'Количество инвестиционных заявок составляет {}'.format(long_term.shape[0])

    num_parts = count_num_parts(long_term)

    if num_parts>0:
        for i in range(num_parts):
            print 'Инвестиционные. Сохраняем часть {}'.format(i+1)
            cut_df = long_term[first:last]
            cut_df = cut_df.drop(['level_0'], axis=1)
            cut_df = cut_df.reset_index(drop=True)
            cut_df = cut_df.reset_index()
            cut_df['index'] =cut_df['index']+1
            cut_shape = cut_df.shape[0]
            cut_shape_all+=cut_shape
            name = u'_инвестиционные_'
            save_cut(cut_df, i, name)
            print 'Инвестиционные. длина сохраненного файла {} записей'.format(cut_shape)
            first+=20
            last+=20
        print 'Инвестиционные. всего сохранили {} записей'.format(cut_shape_all)

def split_save_region (all_df, region):

    """Save shrot-term requests"""

    all_df = read_index()
    try:
        del all_df['Unnamed: 7']
        del all_df['Unnamed: 8']
        del all_df['uid']
    except:
        pass

    short_term = all_df[all_df['level_8']==region]

    cut_shape_all = 0
    first = 0
    last = 20

    print ' '
    print 'Количество заявок составляет {}'.format(short_term.shape[0])

    num_parts = count_num_parts(short_term)

    if num_parts>0:
        for i in range(num_parts):
            print 'Сохраняем часть {}'.format(i+1)
            cut_df = short_term[first:last]
            cut_df = cut_df.drop(['level_0'], axis=1)
            cut_df = cut_df.reset_index(drop=True)
            cut_df = cut_df.reset_index()
            cut_df['index'] =cut_df['index']+1
            cut_shape = cut_df.shape[0]
            cut_shape_all+=cut_shape
            name = u'_по регионам_{}_'.format(region)
            save_cut(cut_df, i, name)
            first+=20
            last+=20

if __name__ == '__main__':
  main()
